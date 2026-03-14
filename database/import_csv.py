"""
Import student data from CSV or Excel files into the database.

Usage:
    python database/import_csv.py                  # import from data/ folder
    python database/import_csv.py --dir path/to/folder  # import from custom folder
    python database/import_csv.py --clear           # clear existing data before import

Expected files in the data folder:
    students.csv     (required)  – student details
    attendance.csv   (optional)  – attendance records
    marks.csv        (optional)  – marks / grades
    cgpa.csv         (optional)  – semester SGPA & CGPA
    backlogs.csv     (optional)  – backlog subjects
    fees.csv         (optional)  – fee details
    exams.csv        (optional)  – upcoming exams
    faculty.csv      (optional)  – faculty details

Files can be .csv or .xlsx (Excel).  For .xlsx, the first sheet is used.
"""

import sys
import os
import csv
import argparse

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from app import create_app
from database.models import (
    db, Student, Attendance, Marks, CGPA, Backlog, Fee, Exam, Faculty
)

# ── Helpers ───────────────────────────────────────────────────────────────────

def read_file(filepath):
    """Read a .csv or .xlsx file and return a list of dicts."""
    ext = os.path.splitext(filepath)[1].lower()

    if ext == ".xlsx":
        try:
            from openpyxl import load_workbook
        except ImportError:
            print("  [!] openpyxl not installed. Run: pip install openpyxl")
            sys.exit(1)
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        if not rows:
            return []
        headers = [str(h).strip() for h in rows[0]]
        return [dict(zip(headers, row)) for row in rows[1:] if any(c is not None for c in row)]

    # Default: CSV
    with open(filepath, newline="", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        return [row for row in reader]


def find_file(folder, base_name):
    """Look for base_name.csv or base_name.xlsx in folder."""
    for ext in (".csv", ".xlsx"):
        path = os.path.join(folder, base_name + ext)
        if os.path.isfile(path):
            return path
    return None


def safe_float(val, default=0.0):
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def safe_int(val, default=0):
    try:
        return int(val)
    except (TypeError, ValueError):
        return default


def strip_val(val):
    """Return stripped string or empty string."""
    return str(val).strip() if val is not None else ""


# ── Import functions ──────────────────────────────────────────────────────────

def import_students(folder):
    path = find_file(folder, "students")
    if not path:
        print("  [!] students.csv/xlsx not found – this file is required!")
        sys.exit(1)

    rows = read_file(path)
    count = 0
    for r in rows:
        reg = strip_val(r.get("reg_number")).upper()
        if not reg:
            continue
        # Skip if already exists
        if Student.query.filter_by(reg_number=reg).first():
            print(f"      Skipping duplicate: {reg}")
            continue
        db.session.add(Student(
            reg_number=reg,
            name=strip_val(r.get("name")),
            branch=strip_val(r.get("branch")),
            year=safe_int(r.get("year"), 1),
            current_semester=safe_int(r.get("current_semester"), 1),
            parent_phone=strip_val(r.get("parent_phone")),
            class_advisor=strip_val(r.get("class_advisor")),
            advisor_email=strip_val(r.get("advisor_email")),
            advisor_phone=strip_val(r.get("advisor_phone")),
        ))
        count += 1
    db.session.flush()
    print(f"  [+] Students: {count} imported")
    return count


def _get_student_map():
    """Return dict mapping reg_number -> student.id"""
    return {s.reg_number: s.id for s in Student.query.all()}


def import_attendance(folder, smap):
    path = find_file(folder, "attendance")
    if not path:
        print("  [-] attendance file not found – skipped")
        return
    rows = read_file(path)
    count = 0
    for r in rows:
        reg = strip_val(r.get("reg_number")).upper()
        sid = smap.get(reg)
        if not sid:
            print(f"      Skipping attendance – student {reg} not found")
            continue
        db.session.add(Attendance(
            student_id=sid,
            subject=strip_val(r.get("subject")),
            total_classes=safe_int(r.get("total_classes")),
            attended=safe_int(r.get("attended")),
            semester=safe_int(r.get("semester")),
        ))
        count += 1
    print(f"  [+] Attendance: {count} records imported")


def import_marks(folder, smap):
    path = find_file(folder, "marks")
    if not path:
        print("  [-] marks file not found – skipped")
        return
    rows = read_file(path)
    count = 0
    for r in rows:
        reg = strip_val(r.get("reg_number")).upper()
        sid = smap.get(reg)
        if not sid:
            print(f"      Skipping marks – student {reg} not found")
            continue
        db.session.add(Marks(
            student_id=sid,
            subject=strip_val(r.get("subject")),
            internal=safe_float(r.get("internal")),
            external=safe_float(r.get("external")),
            total=safe_float(r.get("total")),
            max_marks=safe_float(r.get("max_marks"), 100.0),
            semester=safe_int(r.get("semester")),
        ))
        count += 1
    print(f"  [+] Marks: {count} records imported")


def import_cgpa(folder, smap):
    path = find_file(folder, "cgpa")
    if not path:
        print("  [-] cgpa file not found – skipped")
        return
    rows = read_file(path)
    count = 0
    for r in rows:
        reg = strip_val(r.get("reg_number")).upper()
        sid = smap.get(reg)
        if not sid:
            continue
        db.session.add(CGPA(
            student_id=sid,
            semester=safe_int(r.get("semester")),
            sgpa=safe_float(r.get("sgpa")),
            cgpa=safe_float(r.get("cgpa")),
        ))
        count += 1
    print(f"  [+] CGPA: {count} records imported")


def import_backlogs(folder, smap):
    path = find_file(folder, "backlogs")
    if not path:
        print("  [-] backlogs file not found – skipped")
        return
    rows = read_file(path)
    count = 0
    for r in rows:
        reg = strip_val(r.get("reg_number")).upper()
        sid = smap.get(reg)
        if not sid:
            continue
        db.session.add(Backlog(
            student_id=sid,
            subject=strip_val(r.get("subject")),
            semester=safe_int(r.get("semester")),
            status=strip_val(r.get("status")) or "pending",
        ))
        count += 1
    print(f"  [+] Backlogs: {count} records imported")


def import_fees(folder, smap):
    path = find_file(folder, "fees")
    if not path:
        print("  [-] fees file not found – skipped")
        return
    rows = read_file(path)
    count = 0
    for r in rows:
        reg = strip_val(r.get("reg_number")).upper()
        sid = smap.get(reg)
        if not sid:
            continue
        db.session.add(Fee(
            student_id=sid,
            total_fees=safe_float(r.get("total_fees")),
            paid=safe_float(r.get("paid")),
            scholarship_amount=safe_float(r.get("scholarship_amount")),
            scholarship_name=strip_val(r.get("scholarship_name")) or None,
            payment_history=strip_val(r.get("payment_history")) or "[]",
        ))
        count += 1
    print(f"  [+] Fees: {count} records imported")


def import_exams(folder, smap):
    path = find_file(folder, "exams")
    if not path:
        print("  [-] exams file not found – skipped")
        return
    rows = read_file(path)
    count = 0
    for r in rows:
        reg = strip_val(r.get("reg_number")).upper()
        sid = smap.get(reg)
        if not sid:
            continue
        db.session.add(Exam(
            student_id=sid,
            subject=strip_val(r.get("subject")),
            exam_date=strip_val(r.get("exam_date")),
            exam_type=strip_val(r.get("exam_type")),
            description=strip_val(r.get("description")),
        ))
        count += 1
    print(f"  [+] Exams: {count} records imported")


def import_faculty(folder):
    path = find_file(folder, "faculty")
    if not path:
        print("  [-] faculty file not found – skipped")
        return
    rows = read_file(path)
    count = 0
    for r in rows:
        name = strip_val(r.get("name"))
        if not name:
            continue
        db.session.add(Faculty(
            subject=strip_val(r.get("subject")),
            name=name,
            email=strip_val(r.get("email")),
            phone=strip_val(r.get("phone")),
            department=strip_val(r.get("department")),
        ))
        count += 1
    print(f"  [+] Faculty: {count} records imported")


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Import student data from CSV/Excel files")
    parser.add_argument("--dir", default=os.path.join(os.path.dirname(__file__), "..", "data"),
                        help="Path to folder containing CSV/XLSX files (default: data/)")
    parser.add_argument("--clear", action="store_true",
                        help="Clear ALL existing data before importing")
    args = parser.parse_args()

    data_dir = os.path.abspath(args.dir)
    if not os.path.isdir(data_dir):
        print(f"  [!] Folder not found: {data_dir}")
        sys.exit(1)

    app = create_app()

    with app.app_context():
        if args.clear:
            print("\n  Clearing existing data...")
            db.drop_all()
            db.create_all()
            print("  Done.\n")

        print(f"  Importing from: {data_dir}\n")

        import_students(data_dir)
        smap = _get_student_map()

        import_attendance(data_dir, smap)
        import_marks(data_dir, smap)
        import_cgpa(data_dir, smap)
        import_backlogs(data_dir, smap)
        import_fees(data_dir, smap)
        import_exams(data_dir, smap)
        import_faculty(data_dir)

        db.session.commit()

        total_students = Student.query.count()
        print(f"\n{'='*55}")
        print(f"  Import complete!  Total students in DB: {total_students}")
        print(f"{'='*55}\n")


if __name__ == "__main__":
    main()
